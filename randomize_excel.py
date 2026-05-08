"""
엑셀 파일의 숫자를 연도 제외하고 랜덤으로 바꾸는 스크립트
지원 형식: .xlsx, .xlsm, .xlsb
사용법: python randomize_excel.py <입력파일> [출력파일.xlsx]
※ .xlsb 입력 시 출력은 항상 .xlsx로 저장됩니다.
"""

import sys
import random
import math
import openpyxl


YEAR_MIN = 2004
YEAR_MAX = 2040


def is_year(value) -> bool:
    """정수이고 연도 범위에 해당하면 연도로 판단"""
    return (
        isinstance(value, (int, float))
        and not isinstance(value, bool)
        and value == int(value)
        and YEAR_MIN <= int(value) <= YEAR_MAX
    )


def randomize_number(value):
    """숫자의 자릿수(크기)를 유지하면서 랜덤 값으로 교체"""
    if value == 0:
        return 0

    is_int = isinstance(value, int) or (isinstance(value, float) and value == int(value))
    sign = -1 if value < 0 else 1
    abs_val = abs(value)

    magnitude = 10 ** math.floor(math.log10(abs_val))
    rand_val = random.uniform(magnitude, magnitude * 10)

    if is_int:
        result = sign * round(rand_val)
        return float(result) if isinstance(value, float) else result
    else:
        decimal_places = len(str(abs_val).rstrip('0').split('.')[-1]) if '.' in str(abs_val) else 0
        return sign * round(rand_val, decimal_places)


def process_xlsx(input_path: str, output_path: str) -> tuple[int, int]:
    """.xlsx / .xlsm 처리"""
    wb = openpyxl.load_workbook(input_path)
    changed = 0
    kept_years = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, (int, float)) or isinstance(value, bool):
                    continue
                if is_year(value):
                    kept_years += 1
                    continue
                cell.value = randomize_number(value)
                changed += 1

    wb.save(output_path)
    return changed, kept_years


def process_xlsb(input_path: str, output_path: str) -> tuple[int, int]:
    """.xlsb 처리 — pyxlsb로 읽어서 openpyxl로 .xlsx 저장"""
    try:
        from pyxlsb import open_workbook
    except ImportError:
        print("오류: .xlsb 파일 처리에 pyxlsb 라이브러리가 필요합니다.")
        print("      pip install pyxlsb")
        sys.exit(1)

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)  # 기본 시트 제거

    changed = 0
    kept_years = 0

    with open_workbook(input_path) as wb:
        for sheet_name in wb.sheets:
            ws_out = wb_out.create_sheet(title=sheet_name)
            with wb.get_sheet(sheet_name) as ws:
                for row in ws.rows():
                    for cell in row:
                        value = cell.v  # pyxlsb 셀 값은 .v 속성

                        r, c = cell.r + 1, cell.c + 1  # pyxlsb는 0-based, openpyxl은 1-based

                        if not isinstance(value, (int, float)) or isinstance(value, bool):
                            ws_out.cell(row=r, column=c, value=value)
                            continue

                        if is_year(value):
                            ws_out.cell(row=r, column=c, value=value)
                            kept_years += 1
                        else:
                            ws_out.cell(row=r, column=c, value=randomize_number(value))
                            changed += 1

    wb_out.save(output_path)
    return changed, kept_years


def make_output_path(input_path: str) -> str:
    """출력 파일명 자동 생성 (확장자는 항상 .xlsx)"""
    base = input_path.rsplit('.', 1)[0]
    return f"{base}_randomized.xlsx"


def main():
    if len(sys.argv) < 2:
        print("사용법: python randomize_excel.py <입력파일> [출력파일.xlsx]")
        print("지원 형식: .xlsx, .xlsm, .xlsb")
        sys.exit(1)

    input_path = sys.argv[1]
    ext = input_path.rsplit('.', 1)[-1].lower() if '.' in input_path else ''

    if ext not in ('xlsx', 'xlsm', 'xlsb'):
        print(f"오류: 지원하지 않는 파일 형식입니다 (.{ext})")
        print("지원 형식: .xlsx, .xlsm, .xlsb")
        sys.exit(1)

    if len(sys.argv) >= 3:
        output_path = sys.argv[2]
        # xlsb 입력인데 출력을 xlsb로 지정한 경우 강제로 xlsx로 변경
        if output_path.lower().endswith('.xlsb'):
            output_path = output_path[:-5] + '.xlsx'
            print(f"※ .xlsb 출력은 불가하여 .xlsx로 변경합니다: {output_path}")
    else:
        output_path = make_output_path(input_path)

    print(f"입력 파일: {input_path}")
    print(f"출력 파일: {output_path}")
    print("처리 중...")

    try:
        if ext == 'xlsb':
            changed, kept_years = process_xlsb(input_path, output_path)
        else:
            changed, kept_years = process_xlsx(input_path, output_path)
    except FileNotFoundError:
        print(f"오류: 파일을 찾을 수 없습니다 — {input_path}")
        sys.exit(1)
    except Exception as e:
        print(f"오류 발생: {e}")
        sys.exit(1)

    print("완료!")
    print(f"  변경된 숫자 셀: {changed}개")
    print(f"  유지된 연도 셀: {kept_years}개 ({YEAR_MIN}~{YEAR_MAX} 범위)")


if __name__ == "__main__":
    main()
