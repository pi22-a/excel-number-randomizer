"""
엑셀 '펀드별참고자료' 시트에서 펀드 유형별 데이터를 추출하는 모듈.

사용법:
  # 단독 실행 → 전체 펀드 데이터를 JSON으로 저장
  python extract_excel.py <엑셀파일.xlsx> [출력.json]

  # 모듈로 임포트
  from extract_excel import ExcelExtractor
  ext = ExcelExtractor("파일.xlsx")
  data = ext.read_all()   # dict[펀드유형명, FundData]
  ext.export_json("fund_data.json")
"""

import sys
import json
from dataclasses import dataclass, asdict, field
from typing import Optional


# ── 데이터 구조 ────────────────────────────────────────────────────────────

@dataclass
class Section1Row:
    """섹션1 연도별 행 하나"""
    year: str           # '2004년' ~ '2026년', '합계'
    count: float        # 조합수
    formed: float       # 결성액 (억원)
    moat: float         # 모태출자 (억원)
    moat_ratio: float   # 모태출자비율
    companies: float    # 투자기업수
    invested: float     # 투자금액 (억원)
    per_company: float  # 기업당투자금액 (억원)


@dataclass
class Section2Row:
    """섹션2 조합상태 행 하나"""
    status: str         # '해산완료' / '운용중' / '진행'
    count: float
    formed: float
    moat: float
    moat_ratio: float
    companies: float
    invested: float
    per_company: float


@dataclass
class Section3Row:
    """섹션3 연도별 투자금액·투자여력 행 하나"""
    year: str           # '<=2015' ~ '<=2026', '합계(누적)'
    invested: float
    capacity: float     # 투자여력(연말)
    cumulative: float   # 누적투자금액


@dataclass
class FundData:
    """펀드 유형 하나의 전체 데이터"""
    fund_type: str          # 펀드 유형명 (B1 셀)
    unit: str               # 단위 (B2 셀)

    # 섹션1 요약 문장
    s1_total_count: float = 0       # '05~'26년 총 조합수
    s1_total_formed: float = 0      # 총 결성액
    s1_total_invested: float = 0    # 총 투자액
    s1_year26_count: float = 0      # '26년 선정 조합수
    s1_year26_formed: float = 0     # '26년 선정 결성액
    s1_year26_moat: float = 0       # '26년 선정 모태
    s1_cur_count: float = 0         # 현재(N월) 결성 개수
    s1_cur_formed: float = 0        # 현재(N월) 결성액
    s1_cur_invested: float = 0      # 현재(N월) 투자액
    s1_pending_count: float = 0     # 결성 진행중 조합수

    # 섹션1 연도별 표
    s1_rows: list = field(default_factory=list)    # List[Section1Row]

    # 섹션1 소계
    s1_sub_formed: float = 0        # 조성소계 결성액
    s1_sub_moat: float = 0          # 조성소계 모태출자
    s1_sub_companies: float = 0     # 조성소계 투자기업수
    s1_sub_invested: float = 0      # 조성소계 투자금액
    s1_sub2_formed: float = 0       # 선정및결성소계 결성액
    s1_sub2_moat: float = 0
    s1_sub2_companies: float = 0
    s1_sub2_invested: float = 0

    # 섹션2 조합 상태
    s2_rows: list = field(default_factory=list)    # List[Section2Row]

    # 섹션3 연도별 투자여력
    s3_rows: list = field(default_factory=list)    # List[Section3Row]
    s3_total_invested: float = 0    # 합계(누적) 투자금액


def _float(v) -> float:
    """None이나 문자열을 안전하게 float 변환"""
    if v is None:
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


# ── 시트 파싱 ──────────────────────────────────────────────────────────────

def parse_sheet(ws) -> FundData:
    """
    openpyxl 워크시트 객체에서 FundData를 파싱.
    data_only=True 로 로드한 워크북에서 사용할 것.
    """
    fund_type = _str(ws["B1"].value)
    unit = _str(ws["B2"].value)

    fd = FundData(fund_type=fund_type, unit=unit)

    # ── 섹션1 요약 문장 ─────────────────────────────────────────────────────
    # 행5: '05~'26년 총계
    fd.s1_total_count   = _float(ws.cell(5, 2).value)
    fd.s1_total_formed  = _float(ws.cell(5, 4).value)
    fd.s1_total_invested= _float(ws.cell(5, 6).value)

    # 행6: '26년 선정
    fd.s1_year26_count  = _float(ws.cell(6, 2).value)
    fd.s1_year26_formed = _float(ws.cell(6, 4).value)
    fd.s1_year26_moat   = _float(ws.cell(6, 7).value)

    # 행7: 현재(N월) 결성
    fd.s1_cur_count     = _float(ws.cell(7, 1).value)
    fd.s1_cur_formed    = _float(ws.cell(7, 3).value)
    fd.s1_cur_invested  = _float(ws.cell(7, 5).value)

    # 행8: 결성 진행중
    fd.s1_pending_count = _float(ws.cell(8, 1).value)

    # ── 섹션1 연도별 표 (행11~33: 2004년~2026년, 행34: 합계) ────────────────
    year_rows = []
    for row in range(11, 35):   # 11~34
        year_val = ws.cell(row, 1).value
        if year_val is None:
            continue
        year_rows.append(Section1Row(
            year       = _str(year_val),
            count      = _float(ws.cell(row, 2).value),
            formed     = _float(ws.cell(row, 3).value),
            moat       = _float(ws.cell(row, 4).value),
            moat_ratio = _float(ws.cell(row, 5).value),
            companies  = _float(ws.cell(row, 6).value),
            invested   = _float(ws.cell(row, 7).value),
            per_company= _float(ws.cell(row, 8).value),
        ))
    fd.s1_rows = year_rows

    # ── 섹션1 소계 (행36, 37) ───────────────────────────────────────────────
    fd.s1_sub_formed     = _float(ws.cell(36, 3).value)
    fd.s1_sub_moat       = _float(ws.cell(36, 4).value)
    fd.s1_sub_companies  = _float(ws.cell(36, 5).value)
    fd.s1_sub_invested   = _float(ws.cell(36, 6).value)

    fd.s1_sub2_formed    = _float(ws.cell(37, 3).value)
    fd.s1_sub2_moat      = _float(ws.cell(37, 4).value)
    fd.s1_sub2_companies = _float(ws.cell(37, 5).value)
    fd.s1_sub2_invested  = _float(ws.cell(37, 6).value)

    # ── 섹션2 조합 상태 (행7~9, 열K=11~R=18) ────────────────────────────────
    s2_status = {7: "해산완료", 8: "운용중", 9: "진행"}
    s2_rows = []
    for row, status in s2_status.items():
        s2_rows.append(Section2Row(
            status     = status,
            count      = _float(ws.cell(row, 12).value),
            formed     = _float(ws.cell(row, 13).value),
            moat       = _float(ws.cell(row, 14).value),
            moat_ratio = _float(ws.cell(row, 15).value),
            companies  = _float(ws.cell(row, 16).value),
            invested   = _float(ws.cell(row, 17).value),
            per_company= _float(ws.cell(row, 18).value),
        ))
    fd.s2_rows = s2_rows

    # ── 섹션3 연도별 투자여력 (행7~18: <=2015~<=2026, 행19: 합계) ────────────
    s3_rows = []
    for row in range(7, 20):
        year_val = ws.cell(row, 23).value   # 열W=23
        if year_val is None:
            continue
        s3_rows.append(Section3Row(
            year       = _str(year_val),
            invested   = _float(ws.cell(row, 24).value),
            capacity   = _float(ws.cell(row, 25).value),
            cumulative = _float(ws.cell(row, 26).value),
        ))
    fd.s3_rows = s3_rows

    # 합계(누적) 투자금액은 행19 X열(24)
    fd.s3_total_invested = _float(ws.cell(19, 24).value)

    return fd


# ── 추출기 클래스 ──────────────────────────────────────────────────────────

class ExcelExtractor:
    """
    엑셀 파일에서 펀드 데이터를 추출하는 클래스.

    두 가지 모드:
    1. read_current()  — openpyxl만 사용, 현재 저장된 펀드 유형 1개
    2. read_all()      — xlwings로 Excel 자동화, 전체 48개 유형 순회
    """

    FUND_TYPES = [
        "창업초기", "창업초기(벤처투자조합등)", "창업초기(개인투자조합)",
        "청년창업", "마이크로VC", "재기지원", "4차산업혁명",
        "조선업구조개선", "지방기업", "여성벤처", "소셜임팩트",
        "기술지주", "혁신성장", "스케일업", "M&A", "세컨더리",
        "엔젤세컨더리", "LP지분유동화", "소재부품장비",
        "버팀목", "해외진출",
        "IP·문화산업", "문화일반, FI유치", "수출", "신기술",
        "지역·청년·일자리", "M&A·세컨더리", "아시아문화중심도시육성",
        "한국영화 메인투자", "중저예산한국영화", "한국영화 개봉촉진",
        "관광기업육성", "스포츠산업", "스포츠프로젝트",
        "국토교통혁신", "도시재생", "대학창업", "사회적기업",
        "SaaS", "사이버보안", "메타버스", "공공기술사업화",
        "IP직접투자", "특허기술사업화", "미래환경산업", "보건",
    ]
    # 스마트대한민국은 23년 3월부터 제외

    SHEET_NORMAL = "펀드별참고자료"
    SHEET_LOCAL  = "펀드별참고자료(지방)"  # 지방기업 전용 시트

    def __init__(self, excel_path: str):
        self.excel_path = excel_path

    # ── 모드 1: 현재 값만 읽기 (openpyxl, 빠름) ─────────────────────────────
    def read_current(self) -> FundData:
        """현재 엑셀에 저장된 상태(한 가지 펀드 유형)의 데이터 반환."""
        import openpyxl
        wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        ws = wb[self.SHEET_NORMAL]
        return parse_sheet(ws)

    # ── 모드 2: 전체 유형 순회 (xlwings, Excel 필요) ──────────────────────────
    def read_all(self, show_progress: bool = True) -> dict:
        """
        xlwings로 Excel을 자동 조작하여 모든 펀드 유형 데이터를 반환.
        반환: dict[str, FundData]
        Excel이 설치되어 있어야 합니다.
        """
        try:
            import xlwings as xw
        except ImportError:
            raise RuntimeError("xlwings가 필요합니다: pip install xlwings")

        results = {}
        import openpyxl

        print("Excel 자동화 시작 (백그라운드)...")
        app = xw.App(visible=False, add_book=False)
        try:
            wb_xw = app.books.open(self.excel_path)
            ws_xw = wb_xw.sheets[self.SHEET_NORMAL]

            total = len(self.FUND_TYPES)
            for i, fund_type in enumerate(self.FUND_TYPES, 1):
                if show_progress:
                    bar = '#' * int(30 * i / total) + '-' * (30 - int(30 * i / total))
                    print(f"\r  [{bar}] {i:2d}/{total}  {fund_type[:12]:12s}", end='', flush=True)

                # B1셀에 펀드 유형 입력 → 수식 자동 재계산
                ws_xw["B1"].value = fund_type
                wb_xw.app.calculate()

                # 임시 저장 없이 xlwings로 직접 값 읽기
                fd = self._read_from_xlwings(ws_xw, fund_type)
                results[fund_type] = fd

            print()  # 줄바꿈
            wb_xw.close()
        finally:
            app.quit()

        return results

    def _read_from_xlwings(self, ws_xw, fund_type: str) -> FundData:
        """xlwings 시트 객체에서 FundData 파싱."""
        def v(row, col):
            val = ws_xw.cells(row, col).value
            if val is None:
                return 0.0
            try:
                return float(val)
            except (TypeError, ValueError):
                return 0.0

        def s(row, col):
            val = ws_xw.cells(row, col).value
            return str(val).strip() if val is not None else ""

        fd = FundData(fund_type=fund_type, unit=s(2, 2))

        # 섹션1 요약
        fd.s1_total_count    = v(5, 2)
        fd.s1_total_formed   = v(5, 4)
        fd.s1_total_invested = v(5, 6)
        fd.s1_year26_count   = v(6, 2)
        fd.s1_year26_formed  = v(6, 4)
        fd.s1_year26_moat    = v(6, 7)
        fd.s1_cur_count      = v(7, 1)
        fd.s1_cur_formed     = v(7, 3)
        fd.s1_cur_invested   = v(7, 5)
        fd.s1_pending_count  = v(8, 1)

        # 섹션1 연도별
        year_rows = []
        for row in range(11, 35):
            year_val = ws_xw.cells(row, 1).value
            if year_val is None:
                continue
            year_rows.append(Section1Row(
                year=str(year_val).strip(),
                count=v(row,2), formed=v(row,3), moat=v(row,4),
                moat_ratio=v(row,5), companies=v(row,6),
                invested=v(row,7), per_company=v(row,8),
            ))
        fd.s1_rows = year_rows

        # 섹션1 소계
        fd.s1_sub_formed     = v(36, 3)
        fd.s1_sub_moat       = v(36, 4)
        fd.s1_sub_companies  = v(36, 5)
        fd.s1_sub_invested   = v(36, 6)
        fd.s1_sub2_formed    = v(37, 3)
        fd.s1_sub2_moat      = v(37, 4)
        fd.s1_sub2_companies = v(37, 5)
        fd.s1_sub2_invested  = v(37, 6)

        # 섹션2
        s2_status = {7: "해산완료", 8: "운용중", 9: "진행"}
        s2_rows = []
        for row, status in s2_status.items():
            s2_rows.append(Section2Row(
                status=status,
                count=v(row,12), formed=v(row,13), moat=v(row,14),
                moat_ratio=v(row,15), companies=v(row,16),
                invested=v(row,17), per_company=v(row,18),
            ))
        fd.s2_rows = s2_rows

        # 섹션3
        s3_rows = []
        for row in range(7, 20):
            year_val = ws_xw.cells(row, 23).value
            if year_val is None:
                continue
            s3_rows.append(Section3Row(
                year=str(year_val).strip(),
                invested=v(row,24), capacity=v(row,25), cumulative=v(row,26),
            ))
        fd.s3_rows = s3_rows
        fd.s3_total_invested = v(19, 24)

        return fd

    # ── JSON 내보내기 / 불러오기 ──────────────────────────────────────────────
    def export_json(self, output_path: str, data: dict = None):
        """
        추출한 데이터를 JSON으로 저장.
        data가 None이면 read_all()을 먼저 실행.
        """
        if data is None:
            data = self.read_all()

        serializable = {}
        for key, fd in data.items():
            d = asdict(fd)
            serializable[key] = d

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(serializable, f, ensure_ascii=False, indent=2)
        print(f"저장 완료: {output_path}  ({len(serializable)}개 펀드 유형)")

    @staticmethod
    def load_json(json_path: str) -> dict:
        """JSON에서 데이터 복원. 반환: dict[str, dict]"""
        with open(json_path, encoding='utf-8') as f:
            return json.load(f)


# ── 단독 실행 ──────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("사용법: python extract_excel.py <엑셀파일.xlsx> [출력.json]")
        print()
        print("  전체 펀드 유형을 Excel 자동화로 순회하여 JSON으로 저장합니다.")
        print("  Excel(Office)이 설치되어 있어야 합니다.")
        sys.exit(1)

    excel_path = sys.argv[1]
    json_path  = sys.argv[2] if len(sys.argv) >= 3 else excel_path.rsplit('.', 1)[0] + '_data.json'

    ext = ExcelExtractor(excel_path)

    print(f"입력: {excel_path}")
    print(f"출력: {json_path}")
    print()

    try:
        data = ext.read_all()
        ext.export_json(json_path, data)
    except Exception as e:
        print(f"\n오류: {e}")
        import traceback; traceback.print_exc()
        sys.exit(1)

    print("완료!")


if __name__ == "__main__":
    main()
